from openpyxl import Workbook, load_workbook
import pickle
import os
from tqdm import tqdm
from datetime import datetime
from k_center import read_point_and_process
from side_function import covert_coordinate_from_4326_to_MA

def categorize(file):
    category_dict = {}
    c_k = []

    subcategory_dict = {}
    s_k = []

    severity_dict = {}
    se_k = []

    wb = load_workbook(file)
    sheet = wb.active
    c_max=sheet.max_column

    for i in range(2,sheet.max_row+1):
        cat = sheet.cell(row=i, column=13).value
        s_cat = str(sheet.cell(row=i, column=13).value) + ' - ' + str(sheet.cell(row=i, column=14).value)
        severity=sheet.cell(row=i, column=12).value
        duration=sheet.cell(row=i, column=c_max).value

        if cat not in c_k:
            category_dict[cat] = []
            c_k.append(cat)
        if s_cat not in s_k:
            subcategory_dict[s_cat] = []
            s_k.append(s_cat)

        if severity not in se_k:
            severity_dict[severity] = []
            se_k.append(severity)

        category_dict[cat].append((severity, duration))
        subcategory_dict[s_cat].append((severity, duration))
        severity_dict[severity].append((s_cat, duration))

    filehandler = open('category_dict.obj', 'wb')
    pickle.dump(category_dict, filehandler)
    del category_dict
    filehandler = open('subcategory_dict.obj', 'wb')
    pickle.dump(subcategory_dict, filehandler)
    del subcategory_dict
    filehandler = open('severity_dict.obj', 'wb')
    pickle.dump(severity_dict, filehandler)
    del severity_dict

def write_point_set_and_process(file, duration_range, severity_category, type_list, type_type='subtype'):
    print('reading excel....')
    wb = load_workbook(os.getcwd() + file, data_only=True)
    print('reading done')
    sheet = wb.active
    c_max = sheet.max_column
    point_set=[]
    for i in tqdm(range(2, sheet.max_row + 1)):
        coordinate = sheet.cell(row=i, column=14).value
        if coordinate == None:
            continue
        latitude=coordinate.split()[0]
        longitude=coordinate.split()[1]
        severity = sheet.cell(row=i, column=13).value
        type = sheet.cell(row=i,column=12).value
        subtype = sheet.cell(row=i,column=11).value
        duration = sheet.cell(row=i, column=c_max-1).value
        time1=sheet.cell(row = i, column=2).value
        time1 = time1[:-3] + time1[-2:]
        fmt = '%d-%b-%y %H:%M:%S  %p %z'
        try:
            #ignore the string row in the excel
            start_time = datetime.strptime(time1, fmt)
            if (duration > duration_range[0]) and (duration < duration_range[1]):
                if severity in severity_category:
                    if type_type == 'type':
                        if type in type_list:
                            y_T, x_T = covert_coordinate_from_4326_to_MA((float(latitude), float(longitude)))
                            point_set.append([(x_T, y_T), start_time, duration, severity, type, subtype])
                    elif type_type == 'subtype':
                        if subtype in type_list:
                            y_T, x_T = covert_coordinate_from_4326_to_MA((float(latitude), float(longitude)))
                            point_set.append([(x_T, y_T), start_time, duration, severity, type, subtype])
        except:
            pass
    print('num of filtered incidents', len(point_set))
    filehandler = open('point_set.obj', 'wb')
    pickle.dump(point_set, filehandler)
    read_point_and_process()

def write_in_text(file):
    wb = load_workbook(file, data_only=True)
    sheet = wb.active
    c_max = sheet.max_column

    flink = open('for_QGIS.txt','w+')
    for i in range(2, sheet.max_row + 1):
        cat = sheet.cell(row=i, column=13).value
        s_cat = str(sheet.cell(row=i, column=13).value) + ' - ' + str(sheet.cell(row=i, column=14).value)
        severity = sheet.cell(row=i, column=12).value
        duration = sheet.cell(row=i, column=c_max).value
        coordinate = sheet.cell(row=i, column=21).Kvalue
        latitude=coordinate.split()[0]
        longitude=coordinate.split()[1]
        flink.write('%s,%s,%s,%s,%s,%s\n' % (latitude, longitude, duration, severity, cat, s_cat))
    flink.close()


if __name__ == '__main__':
    #change this to the path of your excel file in your computer
    excel_path = '/MA-ERS_events_2013-2018_0406/MA-ERS_events_2013-2021_correct_Python_TL_930.xlsx'

    #range of duration
    duration_range = (30, 300) #min

    #severity_category
    severity_category = ['Level 2', 'Level 3', 'Level 4']

    #select the type of incidents you want to select for the analysis here
    focus_type = ['Fire', 'Environmental / Hazmat', 'Roadway / Traffic']
    # focus_type = ['Fire']

    # categorize(excel_path)
    # write_in_text(excel_path)
    write_point_set_and_process(excel_path, duration_range, severity_category, focus_type, type_type='type')
