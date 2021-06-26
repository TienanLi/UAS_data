from openpyxl import Workbook, load_workbook
import pickle
from datetime import datetime

def categorize(file):
    category_dict = {}
    c_k = []

    subcategory_dict = {}
    s_k = []

    severity_dict = {}
    se_k = []

    wb = load_workbook(file)
    sheet = wb.active
    c_max=sheet.max_column

    for i in range(2,sheet.max_row+1):
        cat = sheet.cell(row=i, column=13).value
        s_cat = str(sheet.cell(row=i, column=13).value) + ' - ' + str(sheet.cell(row=i, column=14).value)
        severity=sheet.cell(row=i, column=12).value
        duration=sheet.cell(row=i, column=c_max).value

        if cat not in c_k:
            category_dict[cat] = []
            c_k.append(cat)
        if s_cat not in s_k:
            subcategory_dict[s_cat] = []
            s_k.append(s_cat)

        if severity not in se_k:
            severity_dict[severity] = []
            se_k.append(severity)

        category_dict[cat].append((severity, duration))
        subcategory_dict[s_cat].append((severity, duration))
        severity_dict[severity].append((s_cat, duration))

    filehandler = open('category_dict.obj', 'wb')
    pickle.dump(category_dict, filehandler)
    del category_dict
    filehandler = open('subcategory_dict.obj', 'wb')
    pickle.dump(subcategory_dict, filehandler)
    del subcategory_dict
    filehandler = open('severity_dict.obj', 'wb')
    pickle.dump(severity_dict, filehandler)
    del severity_dict

def write_point_set(file):
    wb = load_workbook(file, data_only=True)
    sheet = wb.active
    c_max = sheet.max_column
    point_set=[]
    for i in range(2, sheet.max_row + 1):
        coordinate = sheet.cell(row=i, column=21).value
        latitude=coordinate.split()[0]
        longitude=coordinate.split()[1]
        severity = sheet.cell(row=i, column=12).value
        subtype = sheet.cell(row=i,column=14).value
        duration = sheet.cell(row=i, column=c_max).value
        time1=sheet.cell(row = i, column=2).value
        time1 = time1[:-3] + time1[-2:]
        fmt = '%d-%b-%y %H:%M:%S  %p %z'
        start_time = datetime.strptime(time1, fmt)

        if (duration>30) and (duration<300) and ((severity=='Level 3') or (severity=='Level 2') or (severity=='Level 4')):
            point_set.append(((float(latitude), float(longitude)),start_time,duration,severity,subtype))
    print(len(point_set))
    filehandler = open('point_set.obj', 'wb')
    pickle.dump(point_set, filehandler)


def write_in_text(file):
    wb = load_workbook(file, data_only=True)
    sheet = wb.active
    c_max = sheet.max_column

    flink = open('for_QGIS.txt','w+')
    for i in range(2, sheet.max_row + 1):
        cat = sheet.cell(row=i, column=13).value
        s_cat = str(sheet.cell(row=i, column=13).value) + ' - ' + str(sheet.cell(row=i, column=14).value)
        severity = sheet.cell(row=i, column=12).value
        duration = sheet.cell(row=i, column=c_max).value
        coordinate = sheet.cell(row=i, column=21).value
        latitude=coordinate.split()[0]
        longitude=coordinate.split()[1]
        flink.write('%s,%s,%s,%s,%s,%s\n' % (latitude, longitude, duration, severity, cat, s_cat))
    flink.close()


if __name__ == '__main__':
    # categorize('C:\\Users\\Tienan_Li\\Desktop\\MA-ERS_events_2013-2018_0406\\Processed_by_Python.xlsx')
    # write_in_text('C:\\Users\\Tienan_Li\\Desktop\\Tienan\\projects\\UAS_data\\MA-ERS_events_2013-2018_0406\\Processed_by_Python.xlsx')
    write_point_set('C:\\Users\\Tienan_Li\\Google Drive\\Research_Tienan Li\\2_Projects\\2019_UAS_MassDOT\\MA-ERS_events_2013-2018_0406\\Processed_by_Python.xlsx')